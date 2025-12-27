from pathlib import Path

import numpy as np
import pytest
from openpyxl import load_workbook

from spectro_app.tests.jdx_index_builder_test_utils import load_index_builder


def test_ftir_export_writes_step_sheets(tmp_path: Path) -> None:
    pytest.importorskip("xlsxwriter", exc_type=ImportError)
    module = load_index_builder()
    export_steps = module.export_spectrum_steps_to_xlsx

    x = np.array([4000.0, 3990.0, 3980.0, 3970.0])
    y = np.array([0.1, 0.12, 0.11, 0.09])

    step_registry = [
        {"label": "raw", "x": x, "y": y},
        {"label": "absorbance", "x": x, "y": y + 0.01},
        {"label": "baseline", "x": x, "y": y * 0.9},
        {"label": "baseline_corrected", "x": x, "y": y - 0.02},
        {"label": "smoothed", "x": x, "y": y + 0.005},
        {"label": "peaks_overlay", "x": x, "y": y},
    ]

    output_path = tmp_path / "ftir_steps.xlsx"
    export_steps(step_registry, str(output_path))

    assert output_path.exists(), "Workbook should be written to disk"

    workbook = load_workbook(output_path)
    expected_sheets = {
        "raw",
        "absorbance",
        "baseline",
        "baseline_corrected",
        "smoothed",
        "peaks_overlay",
    }
    assert expected_sheets.issubset(set(workbook.sheetnames))
