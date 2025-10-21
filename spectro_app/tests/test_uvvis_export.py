import csv
import hashlib
import io
import json
import math
import os
from datetime import datetime, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
import pytest

from importlib import metadata
from openpyxl import Workbook, load_workbook

from scipy.signal import savgol_filter

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
try:  # pragma: no cover - optional dependency at test time
    from PyQt6 import QtWidgets
except ImportError:  # pragma: no cover - handled by skip logic
    QtWidgets = None  # type: ignore[assignment]

from spectro_app.app_context import AppContext
from spectro_app.engine import excel_writer
from spectro_app.engine.plugin_api import BatchResult, Spectrum
from spectro_app.plugins.uvvis import pipeline
from spectro_app.plugins.uvvis.plugin import UvVisPlugin

if QtWidgets is not None:  # pragma: no cover - optional UI dependency
    from spectro_app.ui.main_window import MainWindow
else:  # pragma: no cover - optional UI dependency
    MainWindow = None  # type: ignore[assignment]


@pytest.fixture(scope="module")
def qt_app():
    if QtWidgets is None:
        pytest.skip("PyQt6 not available")
    app = QtWidgets.QApplication.instance()
    if app is None:
        app = QtWidgets.QApplication([])
    app.setQuitOnLastWindowClosed(False)
    return app


def _mock_spectrum():
    wl = np.linspace(220.0, 320.0, 401)
    peak1 = 1.2 * np.exp(-0.5 * ((wl - 260.0) / 4.0) ** 2)
    peak2 = 0.8 * np.exp(-0.5 * ((wl - 280.0) / 5.0) ** 2)
    baseline = 0.05 * (wl - wl.min()) / (wl.max() - wl.min())
    intensity = peak1 + peak2 + baseline
    return Spectrum(
        wavelength=wl,
        intensity=intensity,
        meta={"sample_id": "Sample-1", "role": "sample", "mode": "absorbance"},
    )


def _calibration_spectrum(sample_id: str, absorbance: float, *, role: str = "sample") -> Spectrum:
    wl = np.linspace(240.0, 280.0, 201)
    intensity = np.full_like(wl, absorbance, dtype=float)
    return Spectrum(
        wavelength=wl,
        intensity=intensity,
        meta={"sample_id": sample_id, "role": role, "mode": "absorbance", "pathlength_cm": 1.0},
    )


def test_uvvis_analyze_extracts_features():
    plugin = UvVisPlugin()
    spec = _mock_spectrum()
    recipe = {
        "features": {
            "integrals": [
                {"name": "Area_250_280", "min": 250.0, "max": 280.0},
            ],
        }
    }

    processed, qc_rows = plugin.analyze([spec], recipe)

    assert len(processed) == 1
    enriched = processed[0]
    features = enriched.meta.get("features")
    assert features, "features should be captured in spectrum metadata"

    ratios = features.get("band_ratios", {})
    assert "A260_A280" in ratios
    assert np.isfinite(ratios["A260_A280"]["value"])

    integrals = features.get("integrals", {})
    assert "Area_250_280" in integrals
    assert integrals["Area_250_280"] > 0

    peaks = features.get("peaks", [])
    assert peaks and peaks[0]["wavelength"]

    channels = enriched.meta.get("channels", {})
    assert "first_derivative" in channels
    assert channels["first_derivative"].shape == enriched.wavelength.shape

    wl = enriched.wavelength
    diffs = np.diff(wl)
    finite_diffs = diffs[np.isfinite(diffs)]
    if finite_diffs.size:
        delta = float(np.nanmedian(finite_diffs))
        if not np.isfinite(delta) or delta == 0:
            delta = float(np.nanmedian(np.abs(finite_diffs)))
    else:
        delta = 1.0
    if not np.isfinite(delta) or delta == 0:
        delta = 1.0
    delta = abs(delta)
    expected_first = savgol_filter(
        enriched.intensity,
        window_length=5,
        polyorder=2,
        deriv=1,
        delta=delta,
        mode="interp",
    )
    expected_second = savgol_filter(
        enriched.intensity,
        window_length=5,
        polyorder=2,
        deriv=2,
        delta=delta,
        mode="interp",
    )
    np.testing.assert_allclose(channels["first_derivative"], expected_first)
    np.testing.assert_allclose(channels["second_derivative"], expected_second)

    assert len(qc_rows) == 1
    qc_entry = qc_rows[0]
    assert "band_ratios" in qc_entry
    assert "integrals" in qc_entry


def test_uvvis_export_creates_workbook_with_derivatives(tmp_path):
    plugin = UvVisPlugin()
    spec = _mock_spectrum()
    recipe = {
        "export": {"path": str(tmp_path / "uvvis_batch.xlsx")},
        "qc": {"quiet_window": {"min": 240.0, "max": 260.0}},
    }

    processed, qc_rows = plugin.analyze([spec], recipe)
    result = plugin.export(processed, qc_rows, recipe)

    workbook_path = Path(recipe["export"]["path"])
    assert workbook_path.exists(), "Workbook should be written to disk"

    wb = load_workbook(workbook_path)
    expected_sheets = {"Processed_Spectra", "Metadata", "QC_Flags", "Calibration", "Audit_Log"}
    assert expected_sheets.issubset(set(wb.sheetnames))

    ws_processed = wb["Processed_Spectra"]
    header = [cell.value for cell in next(ws_processed.iter_rows(min_row=1, max_row=1))]
    expected_header, expected_rows = excel_writer._processed_table_wide(processed)

    assert header == expected_header
    assert header[0] == "wavelength"
    columns = header[1:]
    assert any(str(name).endswith(":processed") for name in columns)
    assert any(str(name).endswith(":first_derivative") for name in columns)
    assert any(str(name).endswith(":second_derivative") for name in columns)

    data_rows = [list(row) for row in ws_processed.iter_rows(min_row=2, values_only=True)]
    assert len(data_rows) == len(expected_rows)
    for got, expected in zip(data_rows, expected_rows):
        assert len(got) == len(expected)
        for idx, value in enumerate(expected):
            if value is None:
                assert got[idx] is None
            else:
                assert got[idx] == pytest.approx(value)

    ws_metadata = wb["Metadata"]
    metadata_header = [cell.value for cell in next(ws_metadata.iter_rows(min_row=1, max_row=1))]
    assert any(str(value).startswith("features.band_ratios") for value in metadata_header)

    ws_qc = wb["QC_Flags"]
    qc_header = [cell.value for cell in next(ws_qc.iter_rows(min_row=1, max_row=1))]
    assert any("band_ratios" in str(value) for value in qc_header)

    assert result.figures, "Export should include generated plots"
    assert "qc_summary_noise.png" in result.figures
    assert any("Workbook written" in entry for entry in result.audit)
    assert result.report_text, "Narrative report text should be populated"
    for heading in ("Ingestion", "Pre-processing", "Analysis", "Results"):
        assert heading in result.report_text


def test_write_workbook_includes_blank_sheet(tmp_path):
    wavelengths = np.linspace(200.0, 210.0, 6)
    sample_spec = Spectrum(
        wavelength=wavelengths,
        intensity=np.linspace(0.1, 0.6, wavelengths.size),
        meta={"sample_id": "Sample-1", "role": "sample", "mode": "absorbance"},
    )
    blank_spec = Spectrum(
        wavelength=wavelengths,
        intensity=np.zeros_like(wavelengths),
        meta={"sample_id": "Blank-1", "role": "blank", "mode": "absorbance"},
    )

    workbook_path = tmp_path / "with_blanks.xlsx"
    excel_writer.write_workbook(
        str(workbook_path),
        [sample_spec, blank_spec],
        qc_table=[],
        audit=[],
        figures=[],
        calibration=None,
    )

    wb = load_workbook(workbook_path)
    assert "Blanks" in wb.sheetnames

    ws_processed = wb["Processed_Spectra"]
    processed_header = [
        cell.value for cell in next(ws_processed.iter_rows(min_row=1, max_row=1))
    ]
    assert all("Blank-1" not in str(column) for column in processed_header[1:])

    ws_metadata = wb["Metadata"]
    metadata_rows = [list(row) for row in ws_metadata.iter_rows(values_only=True)]
    if len(metadata_rows) > 1:
        sample_rows = metadata_rows[1:]
        assert all(
            "Blank-1" not in {
                str(value) if value is not None else "" for value in row
            }
            for row in sample_rows
        )

    ws_blanks = wb["Blanks"]
    blanks_rows = [list(row) for row in ws_blanks.iter_rows(values_only=True)]
    blank_header = blanks_rows[0]
    assert any("Blank-1" in str(value) for value in blank_header[1:])

    blank_metadata = excel_writer._metadata_rows([blank_spec])
    if blank_metadata:
        metadata_header = sorted({key for row in blank_metadata for key in row.keys()})
        written_header = blanks_rows[-(len(blank_metadata) + 1)]
        assert list(written_header) == metadata_header
        for row in blanks_rows[-len(blank_metadata):]:
            mapped = dict(zip(metadata_header, row))
            assert mapped.get("role") == "blank"
            assert mapped.get("sample_id") == "Blank-1"


def test_export_includes_blanks_by_default(tmp_path):
    plugin = UvVisPlugin()
    wavelengths = np.linspace(200.0, 210.0, 6)
    sample_spec = Spectrum(
        wavelength=wavelengths,
        intensity=np.linspace(0.1, 0.6, wavelengths.size),
        meta={"sample_id": "Sample-1", "role": "sample", "mode": "absorbance"},
    )
    blank_spec = Spectrum(
        wavelength=wavelengths,
        intensity=np.zeros_like(wavelengths),
        meta={"sample_id": "Blank-1", "role": "blank", "mode": "absorbance"},
    )

    recipe = {
        "blank": {"subtract": False, "require": False},
        "export": {"path": str(tmp_path / "include_blanks.xlsx")},
    }

    processed, qc_rows = plugin.analyze([sample_spec, blank_spec], recipe)
    result = plugin.export(processed, qc_rows, recipe)

    wb = load_workbook(recipe["export"]["path"])
    assert "Blanks" in wb.sheetnames
    assert any((spec.meta or {}).get("role") == "blank" for spec in result.processed)
    assert any(row.get("role") == "blank" for row in result.qc_table)


def test_export_excludes_blanks_when_disabled(tmp_path):
    plugin = UvVisPlugin()
    wavelengths = np.linspace(200.0, 210.0, 6)
    sample_spec = Spectrum(
        wavelength=wavelengths,
        intensity=np.linspace(0.1, 0.6, wavelengths.size),
        meta={"sample_id": "Sample-1", "role": "sample", "mode": "absorbance"},
    )
    blank_spec = Spectrum(
        wavelength=wavelengths,
        intensity=np.zeros_like(wavelengths),
        meta={"sample_id": "Blank-1", "role": "blank", "mode": "absorbance"},
    )

    recipe = {
        "blank": {"subtract": False, "require": False},
        "export": {
            "path": str(tmp_path / "exclude_blanks.xlsx"),
            "include_blanks": False,
        },
    }

    processed, qc_rows = plugin.analyze([sample_spec, blank_spec], recipe)
    result = plugin.export(processed, qc_rows, recipe)

    wb = load_workbook(recipe["export"]["path"])
    assert "Blanks" not in wb.sheetnames
    assert all((spec.meta or {}).get("role") != "blank" for spec in result.processed)
    assert all(row.get("role") != "blank" for row in result.qc_table)

def test_uvvis_export_writes_text_summary(tmp_path):
    plugin = UvVisPlugin()
    spec = _mock_spectrum()
    workbook_path = tmp_path / "uvvis_batch.xlsx"
    recipe = {
        "export": {
            "path": str(workbook_path),
            "text_summary_enabled": True,
        }
    }

    processed, qc_rows = plugin.analyze([spec], recipe)
    result = plugin.export(processed, qc_rows, recipe)

    text_path = workbook_path.with_suffix(".txt")
    assert text_path.exists(), "Text summary should be written alongside the workbook"
    contents = text_path.read_text(encoding="utf-8")
    assert result.report_text, "Export should provide narrative text"
    assert contents.strip(), "Text summary file should not be empty"
    assert contents == result.report_text
    for heading in ("Ingestion", "Pre-processing", "Analysis", "Results"):
        assert heading in contents
    assert any("Text summary written" in entry for entry in result.audit)


def test_uvvis_export_generates_noise_histogram_and_trend(tmp_path):
    plugin = UvVisPlugin()
    base = _mock_spectrum()
    base_time = datetime(2024, 1, 1, 9, 0)
    rng = np.random.default_rng(42)
    spectra = []
    for idx in range(6):
        meta = dict(base.meta)
        meta["sample_id"] = f"Sample-{idx + 1}"
        meta["acquired_datetime"] = (base_time + timedelta(minutes=idx * 5)).isoformat()
        noise_scale = 0.002 + 0.001 * idx
        noisy_intensity = base.intensity + rng.normal(scale=noise_scale, size=base.intensity.shape)
        spectra.append(
            Spectrum(
                wavelength=base.wavelength.copy(),
                intensity=noisy_intensity,
                meta=meta,
            )
        )

    recipe = {
        "export": {"path": str(tmp_path / "uvvis_hist_trend.xlsx")},
        "qc": {"quiet_window": {"min": 240.0, "max": 260.0}},
    }
    processed, qc_rows = plugin.analyze(spectra, recipe)
    result = plugin.export(processed, qc_rows, recipe)

    assert "qc_summary_noise.png" in result.figures
    assert "qc_summary_noise_hist.png" in result.figures
    assert "qc_summary_noise_trend.png" in result.figures


def test_uvvis_export_handles_numeric_manifest(tmp_path):
    plugin = UvVisPlugin(enable_manifest=True)
    data_path = tmp_path / "sample.csv"
    df = pd.DataFrame(
        {
            "wavelength": [200.0, 205.0, 210.0],
            "Sample-1": [0.1, 0.2, 0.3],
        }
    )
    df.to_csv(data_path, index=False)

    manifest_path = tmp_path / "experiment_manifest.csv"
    manifest_df = pd.DataFrame(
        {
            "file": [data_path.name],
            "sample_id": ["Sample-1"],
            "pathlength_cm": [np.float64(1.5)],
            "dilution_factor": [np.int64(2)],
        }
    )
    manifest_df.to_csv(manifest_path, index=False)

    spectra = plugin.load([str(manifest_path), str(data_path)])
    assert spectra, "Spectra should be loaded from data file"

    workbook_path = tmp_path / "uvvis_manifest.xlsx"
    recipe_path = tmp_path / "uvvis_manifest.recipe.json"
    recipe = {
        "export": {
            "path": str(workbook_path),
            "recipe_sidecar": str(recipe_path),
        }
    }

    processed, qc_rows = plugin.analyze(spectra, recipe)
    plugin.export(processed, qc_rows, recipe)

    assert workbook_path.exists(), "Workbook should be created"
    assert recipe_path.exists(), "Recipe sidecar should be written"

    first_meta = processed[0].meta
    assert isinstance(first_meta.get("pathlength_cm"), float)
    assert isinstance(first_meta.get("dilution_factor"), int)


def test_uvvis_export_clones_visionlite_sources(tmp_path):
    dsp_text = """sinacsa
Scan
1
Sample01.dsp
nm
400
402
1
3
A
-0.1
0.5
0
Operator Name

2024.05.01. 10:00:00
2024.05.01. 10:05:00
2024.05.01. 09:55:00
0
version 1, 2024.05.01. 10:05:00, Operator Name: |Created; Method: ||

1
0
0
sec
0
0

GAMMA
v7.03 v4.80
143309
2
0
0
0
0
0
0

7
0
0
VISIONlite Scan Version 2.1

#SPECIFIC1
0
0
0.01

#SPECIFIC2

#SPECIFIC3

#DATA
0.1
0.2
0.3
"""
    source_path = tmp_path / "sample01.dsp"
    source_path.write_text(dsp_text, encoding="utf-8")
    original_header = dsp_text.split("#DATA", 1)[0]

    plugin = UvVisPlugin()
    spectra = plugin.load([str(source_path)])
    processed, qc_rows = plugin.analyze(spectra, {})

    assert processed, "Processed spectra should not be empty"
    processed[0].intensity = processed[0].intensity + 0.05

    per_spec_dir = tmp_path / "per_spectrum"
    export_recipe = {
        "export": {
            "path": str(tmp_path / "visionlite.xlsx"),
            "per_spectrum_enabled": True,
            "per_spectrum_format": "original",
            "per_spectrum_directory": str(per_spec_dir),
        }
    }

    plugin.export(processed, qc_rows, export_recipe)

    clone_path = per_spec_dir / f"{source_path.stem}_edited{source_path.suffix}"
    assert clone_path.exists(), "Edited Visionlite clone should be written"

    original_location = source_path.with_name(
        f"{source_path.stem}_edited{source_path.suffix}"
    )
    assert not original_location.exists(), "Clone should be emitted into the override directory"

    clone_text = clone_path.read_text(encoding="utf-8")
    clone_header, clone_data = clone_text.split("#DATA", 1)
    original_lines = original_header.splitlines()
    clone_lines = clone_header.splitlines()
    assert len(clone_lines) == len(original_lines)
    assert len(clone_lines) > 3
    assert clone_lines[3] == clone_path.name
    assert clone_lines[:3] == original_lines[:3]
    assert clone_lines[4:] == original_lines[4:]

    data_lines = [line.strip() for line in clone_data.splitlines() if line.strip() and not line.startswith("#")]
    emitted_values = np.array([float(value) for value in data_lines], dtype=float)
    assert np.allclose(emitted_values, processed[0].intensity)


def test_uvvis_visionlite_clone_preserves_header_grid_after_cropping(tmp_path):
    dsp_text = """sinacsa
Scan
1
Sample02.dsp
nm
400
408
2
5
A
-0.1
0.5
0
Operator Name

2024.05.01. 10:00:00
2024.05.01. 10:05:00
2024.05.01. 09:55:00
0
version 1, 2024.05.01. 10:05:00, Operator Name: |Created; Method: ||

1
0
0
sec
0
0

GAMMA
v7.03 v4.80
143309
2
0
0
0
0
0
0

7
0
0
VISIONlite Scan Version 2.1

#SPECIFIC1
0
0
0.01

#SPECIFIC2

#SPECIFIC3

#DATA
0.1
0.2
0.3
0.4
0.5
"""

    source_path = tmp_path / "sample02.dsp"
    source_path.write_text(dsp_text, encoding="utf-8")
    original_header = dsp_text.split("#DATA", 1)[0]

    plugin = UvVisPlugin()
    spectra = plugin.load([str(source_path)])
    recipe = {"domain": {"min": 402.0, "max": 406.0}}
    processed, qc_rows = plugin.analyze(spectra, recipe)

    assert processed, "Processed spectra should not be empty"
    processed_spec = processed[0]
    processed_spec.wavelength = np.array([402.0, 404.0, 406.0], dtype=float)
    processed_spec.intensity = np.array([0.7, 0.8, 0.9], dtype=float)

    per_spec_dir = tmp_path / "per_spectrum"
    export_recipe = {
        "export": {
            "path": str(tmp_path / "visionlite_cropped.xlsx"),
            "per_spectrum_enabled": True,
            "per_spectrum_format": "original",
            "per_spectrum_directory": str(per_spec_dir),
        }
    }

    plugin.export(processed, qc_rows, export_recipe)

    clone_path = per_spec_dir / f"{source_path.stem}_edited{source_path.suffix}"
    assert clone_path.exists(), "Edited Visionlite clone should be written"

    clone_text = clone_path.read_text(encoding="utf-8")
    clone_header, clone_data = clone_text.split("#DATA", 1)
    original_lines = original_header.splitlines()
    clone_lines = clone_header.splitlines()
    assert len(clone_lines) == len(original_lines)
    assert len(clone_lines) > 3
    assert clone_lines[3] == clone_path.name
    assert clone_lines[:3] == original_lines[:3]
    assert clone_lines[4:] == original_lines[4:]

    data_lines = [line.strip() for line in clone_data.splitlines() if line.strip() and not line.startswith("#")]
    assert len(data_lines) == 5, "Clone should retain header-defined row count"
    emitted_values = np.array([float(value) for value in data_lines], dtype=float)
    expected = np.array([0.1, 0.7, 0.8, 0.9, 0.5], dtype=float)
    assert np.allclose(emitted_values, expected)


def test_uvvis_visionlite_clone_uses_unique_replicate_keys(tmp_path):
    dsp_text = """sinacsa
Scan
1
Sample03.dsp
nm
400
402
1
3
A
-0.1
0.5
0
Operator Name

2024.05.01. 10:00:00
2024.05.01. 10:05:00
2024.05.01. 09:55:00
0
version 1, 2024.05.01. 10:05:00, Operator Name: |Created; Method: ||

1
0
0
sec
0
0

GAMMA
v7.03 v4.80
143309
2
0
0
0
0
0
0

7
0
0
VISIONlite Scan Version 2.1

#SPECIFIC1
0
0
0.01

#SPECIFIC2

#SPECIFIC3

#DATA
0.1
0.2
0.3
"""

    source_path = tmp_path / "sample03.dsp"
    source_path.write_text(dsp_text, encoding="utf-8")

    plugin = UvVisPlugin()
    spectra = plugin.load([str(source_path)])
    processed, qc_rows = plugin.analyze(spectra, {})

    assert processed, "Processed spectra should not be empty"
    processed[0].intensity = processed[0].intensity + 0.02

    export_recipe = {
        "export": {
            "path": str(tmp_path / "visionlite_unique.xlsx"),
            "per_spectrum_enabled": True,
            "per_spectrum_format": "original",
            "per_spectrum_directory": str(tmp_path),
        }
    }

    plugin.export(processed, qc_rows, export_recipe)

    clone_path = tmp_path / f"{source_path.stem}_edited{source_path.suffix}"
    assert clone_path.exists(), "Edited Visionlite clone should be written beside the source"

    reload_plugin = UvVisPlugin()
    reloaded = reload_plugin.load([str(source_path), str(clone_path)])
    assert len(reloaded) == 2, "Both original and clone should reload"

    recipe = {"blank": {"subtract": False, "require": False}}
    preprocessed = reload_plugin.preprocess(reloaded, recipe)
    assert len(preprocessed) == 2

    replicate_keys = {pipeline.replicate_key(spec) for spec in preprocessed}
    assert len(replicate_keys) == 2, "Clone should receive a distinct replicate key"


def test_write_workbook_handles_replicate_channel_arrays(tmp_path):
    wavelengths = np.linspace(220.0, 230.0, 6)
    intensity = np.linspace(0.1, 0.2, wavelengths.size)
    channel_arrays = {
        "raw": np.asarray(intensity + 0.01),
        "first_derivative": np.gradient(intensity),
    }
    replicate_entry = {
        "wavelength": np.asarray(wavelengths),
        "intensity": np.asarray(intensity),
        "meta": {
            "sample_id": "Sample-Rep",
            "channels": channel_arrays,
        },
        "excluded": False,
        "score": np.float64(0.0),
        "leverage": np.float64(0.5),
    }
    spec = Spectrum(
        wavelength=wavelengths,
        intensity=intensity,
        meta={
            "sample_id": "Sample-Rep",
            "role": "sample",
            "mode": "absorbance",
            "replicates": [replicate_entry],
        },
    )

    out_path = tmp_path / "replicates.xlsx"
    result_path = excel_writer.write_workbook(
        str(out_path),
        [spec],
        qc_table=[],
        audit=[],
        figures={},
    )

    workbook_path = Path(result_path)
    assert workbook_path.exists()


def test_uvvis_generate_figures_trend_uses_sanitised_name():
    plugin = UvVisPlugin()
    qc_rows = [
        {
            "sample_id": "QC-1",
            "noise_rsd": 0.8,
            "timestamp": "2024-01-01T09:00:00Z",
        },
        {
            "sample_id": "QC-2",
            "noise_rsd": 1.1,
            "kinetics": {"timestamp": "2024-01-01T09:05:00+00:00"},
        },
    ]

    figures, figure_objs = plugin._generate_figures([], qc_rows)

    trend_keys = [name for name in figures if "noise_trend" in name]
    assert trend_keys, "Noise trend figure should be generated when QC timestamps exist"

    # The sanitised filename should remain stable so downstream exports keep the expected
    # ``qc_summary_noise_trend.png`` path; this guards against regressions where the
    # sanitisation logic was applied multiple times (producing ``..._png.png``).
    assert "qc_summary_noise_trend.png" in figures

    trend_obj_names = [name for name, _ in figure_objs if "noise_trend" in name]
    assert trend_obj_names == ["qc_summary_noise_trend.png"]


def test_uvvis_derivative_respects_recipe_window():
    plugin = UvVisPlugin()
    spec = _mock_spectrum()
    recipe = {
        "smoothing": {"enabled": True, "window": 11, "polyorder": 3},
        "features": {"derivatives": {"enabled": True, "window": 7, "polyorder": 3}},
    }

    processed, _ = plugin.analyze([spec], recipe)

    enriched = processed[0]
    channels = enriched.meta.get("channels", {})
    assert "first_derivative" in channels
    assert "second_derivative" in channels

    wl = enriched.wavelength
    diffs = np.diff(wl)
    finite_diffs = diffs[np.isfinite(diffs)]
    if finite_diffs.size:
        delta = float(np.nanmedian(finite_diffs))
        if not np.isfinite(delta) or delta == 0:
            delta = float(np.nanmedian(np.abs(finite_diffs)))
    else:
        delta = 1.0
    if not np.isfinite(delta) or delta == 0:
        delta = 1.0
    delta = abs(delta)

    expected_first = savgol_filter(
        enriched.intensity,
        window_length=7,
        polyorder=3,
        deriv=1,
        delta=delta,
        mode="interp",
    )
    expected_second = savgol_filter(
        enriched.intensity,
        window_length=7,
        polyorder=3,
        deriv=2,
        delta=delta,
        mode="interp",
    )

    np.testing.assert_allclose(channels["first_derivative"], expected_first)
    np.testing.assert_allclose(channels["second_derivative"], expected_second)


def test_uvvis_export_supports_wide_processed_layout(tmp_path):
    plugin = UvVisPlugin()
    spec_a = _mock_spectrum()
    spec_b = _mock_spectrum()
    spec_b.meta = dict(spec_b.meta)
    spec_b.meta["sample_id"] = "Sample-2"
    recipe = {
        "export": {
            "path": str(tmp_path / "uvvis_wide.xlsx"),
            "processed_layout": "wide",
        }
    }

    processed, qc_rows = plugin.analyze([spec_a, spec_b], recipe)
    plugin.export(processed, qc_rows, recipe)

    workbook_path = Path(recipe["export"]["path"])
    assert workbook_path.exists(), "Wide workbook should be written"

    wb = load_workbook(workbook_path)
    ws_processed = wb["Processed_Spectra"]

    header = [cell.value for cell in next(ws_processed.iter_rows(min_row=1, max_row=1))]
    expected_header, expected_rows = excel_writer._processed_table_wide(processed)

    assert header == expected_header
    assert header[0] == "wavelength"
    assert all(isinstance(name, str) for name in header[1:])

    sample_labels = []
    for idx, spec in enumerate(processed):
        sample_meta = spec.meta or {}
        sample_label_value = excel_writer._clean_value(
            sample_meta.get("sample_id")
            or sample_meta.get("channel")
            or sample_meta.get("blank_id")
            or f"spec_{idx}"
        )
        if sample_label_value in (None, ""):
            sample_label_value = f"spec_{idx}"
        sample_labels.append(str(sample_label_value))

    columns = header[1:]
    stage_to_columns: dict[str, list[str]] = {}
    for column in columns:
        if ":" in column:
            _, stage = column.split(":", 1)
        else:
            stage = ""
        stage_to_columns.setdefault(stage, []).append(column)

    assert "processed" in stage_to_columns
    ordered_stages = list(stage_to_columns.keys())
    assert ordered_stages[0] == "processed"
    for stage, stage_columns in stage_to_columns.items():
        if stage:
            expected = [f"{sample}:{stage}" for sample in sample_labels]
        else:
            expected = sample_labels[:]
        assert stage_columns == expected

    data_rows = [list(row) for row in ws_processed.iter_rows(min_row=2, values_only=True)]
    assert len(data_rows) == len(expected_rows)

    for got, expected in zip(data_rows, expected_rows):
        assert len(got) == len(expected)
        assert got[0] == pytest.approx(expected[0])
        for idx, value in enumerate(expected[1:], start=1):
            if value is None:
                assert got[idx] is None
            else:
                assert got[idx] == pytest.approx(value)


def test_uvvis_export_ui_writes_requested_format(tmp_path, qt_app, monkeypatch):
    if QtWidgets is None or MainWindow is None:
        pytest.skip("PyQt6 not available")
    ctx = AppContext()
    window = MainWindow(ctx)
    try:
        plugin = window._plugin_registry.get("uvvis")
        assert plugin is not None

        spec = _mock_spectrum()
        recipe = {"export": {}}
        processed, qc_rows = plugin.analyze([spec], recipe)

        window._active_plugin = plugin
        window._active_plugin_id = plugin.id
        window._last_result = BatchResult(
            processed=list(processed),
            qc_table=list(qc_rows),
            figures={},
            audit=[],
            report_text=None,
        )
        window._recipe_data.setdefault("export", {})

        export_target = tmp_path / "ui_uvvis_export.xlsx"

        monkeypatch.setattr(
            QtWidgets.QFileDialog,
            "getSaveFileName",
            lambda *args, **kwargs: (str(export_target), "xlsx"),
        )

        monkeypatch.setattr(
            window,
            "_prompt_export_options",
            lambda **kwargs: {
                "per_spectrum_enabled": False,
                "per_spectrum_format": "original",
            },
        )

        window.on_export()

        export_cfg = window._recipe_data.get("export", {})
        assert "processed_layout" not in export_cfg
        assert not export_cfg.get("per_spectrum_enabled")

        assert export_target.exists(), "Workbook should be written via UI export"

        wb = load_workbook(export_target)
        ws_processed = wb["Processed_Spectra"]

        header = [cell.value for cell in next(ws_processed.iter_rows(min_row=1, max_row=1))]
        expected_header, expected_rows = excel_writer._processed_table_wide(processed)
        assert header == expected_header

        data_rows = [list(row) for row in ws_processed.iter_rows(min_row=2, values_only=True)]
        assert data_rows == expected_rows
    finally:
        window.close()


def test_uvvis_export_includes_pipeline_stage_channels(tmp_path):
    plugin = UvVisPlugin()
    wl = np.linspace(400.0, 460.0, 7)
    raw_intensity = np.array([1.0, 1.2, 1.1, 3.5, 3.6, 3.7, 3.8])
    sample = Spectrum(
        wavelength=wl,
        intensity=raw_intensity,
        meta={"sample_id": "StageSample", "role": "sample"},
    )
    blank = Spectrum(
        wavelength=wl,
        intensity=np.full_like(wl, 0.2, dtype=float),
        meta={"role": "blank"},
    )

    coerced = pipeline.coerce_domain(sample, {"min": 400.0, "max": 460.0, "num": 7})
    joined = pipeline.correct_joins(coerced, [3], window=1)
    joined.meta["join_indices"] = (3,)
    despiked = pipeline.despike_spectrum(joined, zscore=2.0, window=3)
    blanked = pipeline.subtract_blank(despiked, blank)
    baseline = pipeline.apply_baseline(blanked, "asls", lam=1e2, p=0.01, niter=10)
    smoothed = pipeline.smooth_spectrum(baseline, window=5, polyorder=2)

    recipe = {"export": {"path": str(tmp_path / "stage_channels.xlsx")}}
    result = plugin.export([smoothed], [{}], recipe)
    processed, qc_rows = plugin.analyze([smoothed], recipe)
    assert qc_rows
    qc_row = qc_rows[0]
    assert "roughness" in qc_row
    assert "raw" in qc_row["roughness"]["channels"]
    assert "smoothed" in qc_row["roughness"]["channels"]
    assert qc_row["roughness"]["channels"]["smoothed"] <= qc_row["roughness"]["channels"]["raw"]
    assert qc_row["roughness_delta"]["raw"] > 0.0

    plugin.export(processed, qc_rows, recipe)

    workbook_path = Path(recipe["export"]["path"])
    assert workbook_path.exists()
    wb = load_workbook(workbook_path)
    ws_processed = wb["Processed_Spectra"]
    header = [cell.value for cell in next(ws_processed.iter_rows(min_row=1, max_row=1))]
    expected_header, expected_rows = excel_writer._processed_table_wide(processed)
    assert header == expected_header

    data_rows = [list(row) for row in ws_processed.iter_rows(min_row=2, values_only=True)]
    assert len(data_rows) == len(expected_rows)
    for got, expected in zip(data_rows, expected_rows):
        assert len(got) == len(expected)
        for idx, value in enumerate(expected):
            if value is None:
                assert got[idx] is None
            else:
                assert got[idx] == pytest.approx(value)

    column_labels = header[1:]
    assert any(str(name).endswith(":raw") for name in column_labels)
    assert any(str(name).endswith(":blanked") for name in column_labels)
    assert any(str(name).endswith(":baseline_corrected") for name in column_labels)
    assert any(str(name).endswith(":joined") for name in column_labels)
    assert any(str(name).endswith(":despiked") for name in column_labels)
    assert any(str(name).endswith(":smoothed") for name in column_labels)
    assert any(name.endswith("join_1.png") for name in result.figures)
    ws_qc = wb["QC_Flags"]
    qc_header = [cell.value for cell in next(ws_qc.iter_rows(min_row=1, max_row=1))]
    assert "roughness.processed" in qc_header
    assert "roughness.channels.raw" in qc_header
    assert "roughness.channels.smoothed" in qc_header
    assert "roughness_delta.raw" in qc_header
    data_row = next(ws_qc.iter_rows(min_row=2, max_row=2, values_only=True))
    processed_idx = qc_header.index("roughness.processed")
    raw_idx = qc_header.index("roughness.channels.raw")
    smoothed_idx = qc_header.index("roughness.channels.smoothed")
    raw_delta_idx = qc_header.index("roughness_delta.raw")
    assert data_row[processed_idx] is not None
    assert data_row[raw_idx] is not None
    assert data_row[smoothed_idx] is not None
    assert data_row[raw_idx] > data_row[smoothed_idx]
    assert data_row[raw_delta_idx] > 0.0


def test_uvvis_export_emits_svg_when_requested(tmp_path):
    plugin = UvVisPlugin()
    spec = _mock_spectrum()
    recipe = {
        "export": {
            "path": str(tmp_path / "uvvis_svg.xlsx"),
            "formats": ["png", "svg"],
        }
    }

    processed, qc_rows = plugin.analyze([spec], recipe)
    result = plugin.export(processed, qc_rows, recipe)

    expected_base = Path("Sample-1_processed")
    png_name = expected_base.with_suffix(".png").name
    svg_name = expected_base.with_suffix(".svg").name
    assert png_name in result.figures
    assert svg_name in result.figures
    svg_stems = {Path(name).stem for name in result.figures if name.endswith(".svg")}
    png_stems = {Path(name).stem for name in result.figures if name.endswith(".png")}
    assert svg_stems <= png_stems


def test_uvvis_export_recipe_sidecar_sanitises_nested_arrays(tmp_path):
    plugin = UvVisPlugin()
    spec = _mock_spectrum()
    nested = np.array(
        [np.array([1.0, 2.0]), np.array([3.0, 4.0])], dtype=object
    )
    recipe = {
        "export": {
            "path": str(tmp_path / "uvvis_nested.xlsx"),
            "recipe_path": str(tmp_path / "uvvis_nested.recipe.json"),
        },
        "params": {"nested": nested},
    }

    processed, qc_rows = plugin.analyze([spec], recipe)
    plugin.export(processed, qc_rows, recipe)

    sidecar_path = Path(recipe["export"]["recipe_path"])
    assert sidecar_path.exists(), "Recipe sidecar should be written"
    data = json.loads(sidecar_path.read_text())
    assert data["params"]["nested"] == [[1.0, 2.0], [3.0, 4.0]]
def test_uvvis_export_audit_includes_runtime_and_input_hash(tmp_path):
    plugin = UvVisPlugin()
    spec = _mock_spectrum()
    source_path = tmp_path / "sample_source.csv"
    source_bytes = b"wavelength,intensity\n1,2\n"
    source_path.write_bytes(source_bytes)
    spec.meta["source_file"] = str(source_path)
    recipe = {"export": {"path": str(tmp_path / "audit.xlsx")}}

    processed, qc_rows = plugin.analyze([spec], recipe)
    audit_entries = plugin._build_audit_entries(processed, qc_rows, recipe, figures={})

    package_label = "spectro-app"
    version = "unknown"
    for candidate in ("spectro-app", "spectro_app"):
        try:
            version = metadata.version(candidate)
        except metadata.PackageNotFoundError:
            continue
        else:
            package_label = candidate
            break
    runtime_token = f"{package_label}=={version}"
    assert any(runtime_token in entry for entry in audit_entries)

    digest = hashlib.sha256(source_bytes).hexdigest()
    assert any(digest in entry for entry in audit_entries)


def test_uvvis_audit_distinguishes_source_file_digests(tmp_path):
    plugin = UvVisPlugin()
    spec_a = _mock_spectrum()
    spec_b = _mock_spectrum()
    spec_b.meta = dict(spec_b.meta)
    spec_b.meta["sample_id"] = "Sample-2"

    file_a = tmp_path / "source_a.csv"
    file_b = tmp_path / "source_b.csv"
    bytes_a = b"alpha,1\n"
    bytes_b = b"beta,2\n"
    file_a.write_bytes(bytes_a)
    file_b.write_bytes(bytes_b)
    spec_a.meta["source_file"] = str(file_a)
    spec_b.meta["source_file"] = str(file_b)

    recipe = {"export": {"path": str(tmp_path / "audit.xlsx")}}

    processed, qc_rows = plugin.analyze([spec_a, spec_b], recipe)
    audit_entries = plugin._build_audit_entries(processed, qc_rows, recipe, figures={})

    digest_a = hashlib.sha256(bytes_a).hexdigest()
    digest_b = hashlib.sha256(bytes_b).hexdigest()

    joined = "\n".join(audit_entries)
    assert f"sha256:{digest_a}" in joined
    assert f"sha256:{digest_b}" in joined


def test_uvvis_audit_reports_missing_source_file(tmp_path):
    plugin = UvVisPlugin()
    spec = _mock_spectrum()
    missing_path = tmp_path / "missing_source.csv"
    spec.meta["source_file"] = str(missing_path)
    recipe = {"export": {"path": str(tmp_path / "audit.xlsx")}}

    processed, qc_rows = plugin.analyze([spec], recipe)
    audit_entries = plugin._build_audit_entries(processed, qc_rows, recipe, figures={})

    joined = "\n".join(audit_entries)
    assert "source_hash_error=file_not_found" in joined
def test_uvvis_export_with_workbook_path_creates_sidecar_and_pdf(tmp_path):
    plugin = UvVisPlugin()
    spec = _mock_spectrum()
    recipe = {
        "export": {
            "path": str(tmp_path / "uvvis_batch.xlsx"),
            "recipe_sidecar": True,
            "pdf_report": True,
        }
    }

    processed, qc_rows = plugin.analyze([spec], recipe)

    result = plugin.export(processed, qc_rows, recipe)

    workbook_path = tmp_path / "uvvis_batch.xlsx"
    recipe_sidecar = tmp_path / "uvvis_batch.recipe.json"
    pdf_report = tmp_path / "uvvis_batch.pdf"

    assert workbook_path.exists()
    assert recipe_sidecar.exists()
    assert pdf_report.exists()

    audit_text = "\n".join(result.audit)
    assert "Workbook written" in audit_text
    assert "Recipe sidecar written" in audit_text
    assert "PDF report written" in audit_text


def test_uvvis_calibration_success(tmp_path):
    plugin = UvVisPlugin()
    slope = 0.12
    intercept = 0.02
    standards = [("Std-0", 0.0), ("Std-1", 5.0), ("Std-2", 10.0)]
    unknowns = [("Sample-1", 7.5)]

    spectra = [
        _calibration_spectrum(sample_id, intercept + slope * conc, role="standard")
        for sample_id, conc in standards
    ]
    spectra.extend(
        _calibration_spectrum(sample_id, intercept + slope * conc, role="sample")
        for sample_id, conc in unknowns
    )

    recipe = {
        "calibration": {
            "targets": [
                {
                    "name": "Analyte",
                    "wavelength": 260.0,
                    "standards": [
                        {"sample_id": sample_id, "concentration": conc}
                        for sample_id, conc in standards
                    ],
                    "unknowns": [{"sample_id": sample_id} for sample_id, _ in unknowns],
                }
            ],
        },
        "export": {"path": str(tmp_path / "calibration_ok.xlsx")},
    }

    processed, qc_rows = plugin.analyze(spectra, recipe)
    results = plugin.last_calibration_results

    assert results and results["status"] == "ok"
    target = results["targets"][0]
    assert pytest.approx(target["fit"]["slope"], rel=1e-6) == slope
    unknown_entry = next(entry for entry in target["unknowns"] if entry["sample_id"] == "Sample-1")
    assert unknown_entry["status"] == "ok"
    assert unknown_entry["predicted_concentration"] == pytest.approx(7.5, rel=1e-4)
    assert unknown_entry["predicted_concentration_std_err"] == pytest.approx(0.0, abs=1e-12)

    qc_unknown = next(row for row in qc_rows if row["sample_id"] == "Sample-1")
    qc_calibration = qc_unknown.get("calibration", {}).get("Analyte")
    assert qc_calibration
    assert qc_calibration["predicted_concentration"] == pytest.approx(7.5, rel=1e-4)
    assert qc_calibration["predicted_concentration_std_err"] == pytest.approx(0.0, abs=1e-12)

    result = plugin.export(processed, qc_rows, recipe)
    workbook_path = Path(recipe["export"]["path"])
    assert workbook_path.exists()
    wb = load_workbook(workbook_path)
    ws_calibration = wb["Calibration"]
    rows = list(ws_calibration.iter_rows(values_only=True))
    summary_header = rows[0]
    assert summary_header[:3] == ("Target", "Status", "Slope")
    analyte_summary = next(row for row in rows if row and row[0] == "Analyte")
    assert analyte_summary[1] == "ok"
    assert analyte_summary[2] == pytest.approx(slope, rel=1e-6)

    standards_header = next(row for row in rows if row and "Response (A/cm)" in row)
    assert standards_header[:3] == ("Sample ID", "Concentration", "Response (A/cm)")
    assert "Applied Weight" in standards_header
    applied_weight_idx = standards_header.index("Applied Weight")
    std_row = next(row for row in rows if row and row[0] == "Std-1")
    assert std_row[2] == pytest.approx(intercept + slope * 5.0, rel=1e-6)
    assert std_row[applied_weight_idx] == pytest.approx(1.0)

    unknowns_header = next(row for row in rows if row and "Predicted Concentration" in row)
    assert unknowns_header[0] == "Sample ID"
    assert "Predicted Concentration StdErr" in unknowns_header
    pred_idx = unknowns_header.index("Predicted Concentration")
    se_idx = unknowns_header.index("Predicted Concentration StdErr")
    sample_row = next(row for row in rows if row and row[0] == "Sample-1")
    assert sample_row[pred_idx] == pytest.approx(7.5, rel=1e-4)
    assert sample_row[se_idx] == pytest.approx(0.0, abs=1e-12)
    assert any("Calibration Analyte" in entry and "conc_se=" in entry for entry in result.audit)
    assert any(name.startswith("calibration_Analyte") for name in result.figures)


def test_uvvis_calibration_weighted_regression(tmp_path):
    plugin = UvVisPlugin()
    standards = [
        ("Std-0", 0.0, 0.05, 1.0),
        ("Std-1", 5.0, 0.65, 1.0),
        ("Std-2", 10.0, 1.05, 5.0),
    ]
    concentrations = np.array([item[1] for item in standards], dtype=float)
    responses = np.array([item[2] for item in standards], dtype=float)
    weights = np.array([item[3] for item in standards], dtype=float)
    coeff = np.polyfit(concentrations, responses, 1, w=weights)
    expected_slope, expected_intercept = coeff
    target_conc = 6.0
    unknown_response = expected_intercept + expected_slope * target_conc

    spectra = [
        _calibration_spectrum(sample_id, response, role="standard")
        for sample_id, _, response, _ in standards
    ]
    spectra.append(
        _calibration_spectrum("Sample-1", unknown_response, role="sample")
    )

    recipe = {
        "calibration": {
            "targets": [
                {
                    "name": "Analyte",
                    "wavelength": 260.0,
                    "standards": [
                        {
                            "sample_id": sample_id,
                            "concentration": conc,
                            "weight": weight,
                        }
                        for sample_id, conc, _, weight in standards
                    ],
                    "unknowns": [{"sample_id": "Sample-1"}],
                }
            ],
        },
        "export": {"path": str(tmp_path / "calibration_weighted.xlsx")},
    }

    processed, qc_rows = plugin.analyze(spectra, recipe)
    result = plugin.export(processed, qc_rows, recipe)

    calibration = plugin.last_calibration_results
    assert calibration and calibration["status"] == "ok"
    target = calibration["targets"][0]
    fit = target["fit"]
    assert fit["weighting_applied"] is True
    assert fit["weights"]["Std-0"] == pytest.approx(1.0)
    assert fit["weights"]["Std-1"] == pytest.approx(1.0)
    assert fit["weights"]["Std-2"] == pytest.approx(5.0)
    assert fit["slope"] == pytest.approx(expected_slope, rel=1e-6)
    assert fit["intercept"] == pytest.approx(expected_intercept, rel=1e-6)
    conc_sigma = fit["concentration_std_per_response"]
    assert math.isfinite(conc_sigma) and conc_sigma > 0

    unknown_entry = target["unknowns"][0]
    assert unknown_entry["predicted_concentration"] == pytest.approx(target_conc, rel=1e-6)
    assert unknown_entry["predicted_concentration_std_err"] == pytest.approx(conc_sigma, rel=1e-6)

    std_entries = {entry["sample_id"]: entry for entry in target["standards"]}
    assert std_entries["Std-2"]["applied_weight"] == pytest.approx(5.0)

    qc_standard = next(row for row in qc_rows if row["sample_id"] == "Std-2")
    qc_calibration = qc_standard.get("calibration", {}).get("Analyte")
    assert qc_calibration["applied_weight"] == pytest.approx(5.0)

    workbook_path = Path(recipe["export"]["path"])
    wb = load_workbook(workbook_path)
    ws_calibration = wb["Calibration"]
    rows = list(ws_calibration.iter_rows(values_only=True))
    standards_header = next(row for row in rows if row and "Applied Weight" in row)
    applied_idx = standards_header.index("Applied Weight")
    configured_idx = standards_header.index("Configured Weight")
    std_row = next(row for row in rows if row and row[0] == "Std-2")
    assert std_row[configured_idx] == pytest.approx(5.0)
    assert std_row[applied_idx] == pytest.approx(5.0)

    assert any("weighting=weighted" in entry for entry in result.audit)


def test_uvvis_calibration_uncertainty_propagates(tmp_path):
    plugin = UvVisPlugin()
    slope = 0.11
    intercept = 0.04
    standards = [
        ("Std-0", 0.0, intercept + slope * 0.0 + 0.005),
        ("Std-1", 5.0, intercept + slope * 5.0 - 0.008),
        ("Std-2", 10.0, intercept + slope * 10.0 + 0.012),
        ("Std-3", 15.0, intercept + slope * 15.0 - 0.01),
    ]
    unknown_sample = ("Sample-1", 7.5, intercept + slope * 7.5)

    spectra = [
        _calibration_spectrum(sample_id, response, role="standard")
        for sample_id, _, response in standards
    ]
    spectra.append(_calibration_spectrum(unknown_sample[0], unknown_sample[2], role="sample"))

    recipe = {
        "calibration": {
            "targets": [
                {
                    "name": "Analyte",
                    "wavelength": 260.0,
                    "standards": [
                        {"sample_id": sample_id, "concentration": conc}
                        for sample_id, conc, _ in standards
                    ],
                    "unknowns": [{"sample_id": unknown_sample[0]}],
                }
            ],
        },
        "export": {"path": str(tmp_path / "calibration_uncertainty.xlsx")},
    }

    processed, qc_rows = plugin.analyze(spectra, recipe)
    result = plugin.export(processed, qc_rows, recipe)

    calibration = plugin.last_calibration_results
    assert calibration and calibration["status"] == "ok"
    target = calibration["targets"][0]
    fit = target["fit"]
    conc_sigma = fit["concentration_std_per_response"]
    assert math.isfinite(conc_sigma) and conc_sigma > 0

    unknown_entry = target["unknowns"][0]
    assert unknown_entry["predicted_concentration"] == pytest.approx(unknown_sample[1], rel=1e-3)
    assert unknown_entry["predicted_concentration_std_err"] == pytest.approx(conc_sigma, rel=1e-6)

    qc_unknown = next(row for row in qc_rows if row["sample_id"] == unknown_sample[0])
    qc_calibration = qc_unknown["calibration"]["Analyte"]
    assert qc_calibration["predicted_concentration_std_err"] == pytest.approx(conc_sigma, rel=1e-6)

    processed_meta = next(spec.meta for spec in processed if spec.meta.get("sample_id") == unknown_sample[0])
    meta_calibration = processed_meta["calibration"]["Analyte"]
    assert meta_calibration["predicted_concentration_std_err"] == pytest.approx(conc_sigma, rel=1e-6)

    workbook_path = Path(recipe["export"]["path"])
    wb = load_workbook(workbook_path)
    ws_calibration = wb["Calibration"]
    rows = list(ws_calibration.iter_rows(values_only=True))
    unknowns_header = next(row for row in rows if row and "Predicted Concentration" in row)
    pred_idx = unknowns_header.index("Predicted Concentration")
    se_idx = unknowns_header.index("Predicted Concentration StdErr")
    sample_row = next(row for row in rows if row and row[0] == unknown_sample[0])
    assert sample_row[pred_idx] == pytest.approx(unknown_entry["predicted_concentration"], rel=1e-6)
    assert sample_row[se_idx] == pytest.approx(conc_sigma, rel=1e-6)

    audit_text = "\n".join(result.audit)
    assert "conc_se=" in audit_text

def test_uvvis_calibration_insufficient_standards():
    plugin = UvVisPlugin()
    spectra = [
        _calibration_spectrum("Std-0", 0.05, role="standard"),
        _calibration_spectrum("Sample-1", 0.18, role="sample"),
    ]

    recipe = {
        "calibration": {
            "targets": [
                {
                    "name": "Analyte",
                    "wavelength": 260.0,
                    "standards": [{"sample_id": "Std-0", "concentration": 0.0}],
                    "unknowns": [{"sample_id": "Sample-1"}],
                }
            ]
        }
    }

    processed, qc_rows = plugin.analyze(spectra, recipe)
    results = plugin.last_calibration_results
    assert results and results["targets"][0]["status"] == "failed"
    errors = " ".join(results["targets"][0]["errors"])
    assert "At least two" in errors

    qc_unknown = next(row for row in qc_rows if row["sample_id"] == "Sample-1")
    calibration_meta = qc_unknown.get("calibration", {}).get("Analyte")
    assert calibration_meta["status"] == "no_model"


def test_uvvis_calibration_poor_regression():
    plugin = UvVisPlugin()
    standards = [
        ("Std-0", 0.0, 0.02),
        ("Std-1", 5.0, 0.6),
        ("Std-2", 10.0, 0.25),
    ]
    unknown = ("Sample-1", 7.0, 0.31)

    spectra = [
        _calibration_spectrum(sample_id, absorbance, role="standard")
        for sample_id, _, absorbance in standards
    ]
    spectra.append(_calibration_spectrum(unknown[0], unknown[2], role="sample"))

    recipe = {
        "calibration": {
            "r2_threshold": 0.995,
            "targets": [
                {
                    "name": "Analyte",
                    "wavelength": 260.0,
                    "standards": [
                        {"sample_id": sample_id, "concentration": conc}
                        for sample_id, conc, _ in standards
                    ],
                    "unknowns": [{"sample_id": unknown[0]}],
                }
            ],
        }
    }

    _, qc_rows = plugin.analyze(spectra, recipe)
    results = plugin.last_calibration_results
    assert results and results["targets"][0]["status"] == "failed"
    assert any("R^2" in error for error in results["targets"][0]["errors"])

    qc_unknown = next(row for row in qc_rows if row["sample_id"] == unknown[0])
    calibration_meta = qc_unknown.get("calibration", {}).get("Analyte")
    assert calibration_meta["status"] == "no_model"
def test_uvvis_export_writes_sidecar_and_pdf(tmp_path):
    plugin = UvVisPlugin()
    spec = _mock_spectrum()
    workbook = tmp_path / "uvvis_batch.xlsx"
    recipe_path = tmp_path / "uvvis_batch.recipe.json"
    pdf_path = tmp_path / "uvvis_batch.pdf"
    recipe = {
        "export": {
            "path": str(workbook),
            "recipe_path": str(recipe_path),
            "pdf_path": str(pdf_path),
        },
        "features": {"integrals": [{"name": "Area", "min": 250.0, "max": 270.0}]},
    }

    processed, qc_rows = plugin.analyze([spec], recipe)
    result = plugin.export(processed, qc_rows, recipe)

    assert workbook.exists(), "Workbook should be generated"
    assert recipe_path.exists(), "Recipe sidecar should be generated"
    assert pdf_path.exists(), "PDF report should be generated"
    assert pdf_path.stat().st_size > 0

    exported_recipe = json.loads(recipe_path.read_text(encoding="utf-8"))
    assert "features" in exported_recipe
    assert any("Recipe sidecar written" in entry for entry in result.audit)
    assert any("PDF report written" in entry for entry in result.audit)


def test_uvvis_export_clone_sources_preserves_headers(tmp_path):
    plugin = UvVisPlugin()

    csv_path = tmp_path / "clone_source.csv"
    csv_lines = [
        "#Instrument: DemoCSV",
        "#Operator: QA",
        "Wavelength,SampleA,SampleB",
        "200,0.1,0.2",
        "201,0.2,0.3",
        "202,0.3,0.4",
    ]
    csv_path.write_text("\n".join(csv_lines) + "\n", encoding="utf-8")

    excel_path = tmp_path / "clone_source.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Instrument", "DemoXLS"])
    ws.append(["Operator", "QA"])
    ws.append(["Wavelength", "SampleC", "SampleD"])
    ws.append([200.0, 0.5, 0.7])
    ws.append([201.0, 0.6, 0.8])
    wb.save(excel_path)

    spectra = plugin.load([str(csv_path), str(excel_path)])

    processed: list[Spectrum] = []
    headers_by_source: dict[Path, tuple[str, ...]] = {}
    processed_map: dict[Path, dict[str, np.ndarray]] = {}
    for spec in spectra:
        meta = dict(spec.meta)
        source = Path(meta["source_file"])
        raw_header = tuple(str(line) for line in meta.get("_raw_header_lines", ()))
        headers_by_source.setdefault(source, raw_header)
        column = str(meta.get("channel") or meta.get("sample_id"))
        processed_spec = Spectrum(
            wavelength=spec.wavelength.copy(),
            intensity=spec.intensity + 1.0,
            meta=meta,
        )
        processed.append(processed_spec)
        processed_map.setdefault(source, {})[column] = processed_spec.intensity

    recipe = {
        "export": {
            "path": str(tmp_path / "clone_batch.xlsx"),
            "clone_sources": True,
        }
    }

    result = plugin.export(processed, [], recipe)

    csv_clone = csv_path.with_name(f"{csv_path.stem}_edited{csv_path.suffix}")
    excel_clone = excel_path.with_name(f"{excel_path.stem}_edited{excel_path.suffix}")

    assert csv_clone.exists()
    assert excel_clone.exists()

    csv_header = list(headers_by_source.get(csv_path, ()))
    csv_output_lines = csv_clone.read_text(encoding="utf-8").splitlines()
    if csv_header:
        assert csv_output_lines[: len(csv_header)] == csv_header
    csv_data = "\n".join(csv_output_lines[len(csv_header) :])
    csv_df = pd.read_csv(io.StringIO(csv_data))
    for column, expected in processed_map[csv_path].items():
        np.testing.assert_allclose(csv_df[column].to_numpy(dtype=float), expected)

    wb_clone = load_workbook(excel_clone, data_only=True)
    ws_clone = wb_clone.active
    rows = list(ws_clone.values)
    excel_header = list(headers_by_source.get(excel_path, ()))
    if excel_header:
        observed = [row[0] for row in rows[: len(excel_header)]]
        assert observed == excel_header
    data_rows = rows[len(excel_header) :]
    header_row = data_rows[0]
    excel_df = pd.DataFrame(data_rows[1:], columns=header_row)
    for column, expected in processed_map[excel_path].items():
        np.testing.assert_allclose(
            np.asarray(excel_df[column], dtype=float), expected
        )

    csv_message = f"Edited source copy written to {csv_clone}"
    excel_message = f"Edited source copy written to {excel_clone}"
    assert csv_message in result.audit
    assert excel_message in result.audit


def test_clean_value_sanitises_formula_strings(tmp_path):
    plugin = UvVisPlugin()
    spec = _mock_spectrum()
    spec.meta["sample_id"] = "=2+2"
    recipe = {"export": {"path": str(tmp_path / "uvvis_batch.xlsx")}}

    processed, qc_rows = plugin.analyze([spec], recipe)
    plugin.export(processed, qc_rows, recipe)

    wb = load_workbook(recipe["export"]["path"])
    ws_processed = wb["Processed_Spectra"]
    header = [cell.value for cell in next(ws_processed.iter_rows(min_row=1, max_row=1))]
    expected_header, expected_rows = excel_writer._processed_table_wide(processed)
    assert header == expected_header
    assert any(str(name).startswith("'=2+2:") for name in header[1:])

    data_rows = [list(row) for row in ws_processed.iter_rows(min_row=2, values_only=True)]
    assert len(data_rows) == len(expected_rows)
    for got, expected in zip(data_rows, expected_rows):
        assert len(got) == len(expected)
        for idx, value in enumerate(expected):
            if value is None:
                assert got[idx] is None
            else:
                assert got[idx] == pytest.approx(value)


def test_uvvis_export_per_spectrum_original_header(tmp_path):
    plugin = UvVisPlugin()
    source_dir = tmp_path / "source"
    source_dir.mkdir()
    source_path = source_dir / "input.csv"
    source_path.write_text("# header\nwavelength,SampleA\n200,0.1\n201,0.2\n")
    spectrum = Spectrum(
        wavelength=np.array([200.0, 201.0], dtype=float),
        intensity=np.array([0.3, 0.4], dtype=float),
        meta={
            "sample_id": "SampleA",
            "channel": "SampleA",
            "role": "sample",
            "mode": "absorbance",
            "source_file": str(source_path),
            "_source_delimiter": ",",
            "_source_decimal": ".",
            "_raw_header_lines": ["# header"],
        },
    )
    recipe = {
        "export": {
            "path": str(tmp_path / "batch.xlsx"),
            "per_spectrum_enabled": True,
            "per_spectrum_format": "original",
        }
    }

    processed, qc_rows = plugin.analyze([spectrum], recipe)
    result = plugin.export(processed, qc_rows, recipe)

    workbook_dir = Path(recipe["export"]["path"]).parent
    clone_path = workbook_dir / "input_edited.csv"
    assert clone_path.exists()
    assert not (source_dir / "input_edited.csv").exists()
    report_results = plugin._report_context.get("results", {})
    per_spec_paths = report_results.get("per_spectrum_exports") or []
    assert str(clone_path) in per_spec_paths
    assert any(str(clone_path) in entry for entry in result.audit)


def test_uvvis_export_per_spectrum_xlsx(tmp_path):
    plugin = UvVisPlugin()
    spectrum = _mock_spectrum()
    recipe = {
        "export": {
            "path": str(tmp_path / "uvvis_batch.xlsx"),
            "per_spectrum_enabled": True,
            "per_spectrum_format": "xlsx",
        }
    }

    processed, qc_rows = plugin.analyze([spectrum], recipe)
    result = plugin.export(processed, qc_rows, recipe)

    emitted_path = tmp_path / "Sample-1_1.xlsx"
    assert emitted_path.exists()
    report_results = plugin._report_context.get("results", {})
    per_spec_paths = report_results.get("per_spectrum_exports") or []
    assert str(emitted_path) in per_spec_paths
    assert any(str(emitted_path) in entry for entry in result.audit)


def test_uvvis_export_per_spectrum_csv(tmp_path):
    plugin = UvVisPlugin()
    spectrum = _mock_spectrum()
    recipe = {
        "export": {
            "path": str(tmp_path / "uvvis_batch.xlsx"),
            "per_spectrum_enabled": True,
            "per_spectrum_format": "csv",
        }
    }

    processed, qc_rows = plugin.analyze([spectrum], recipe)
    result = plugin.export(processed, qc_rows, recipe)

    emitted_path = tmp_path / "Sample-1_1.csv"
    assert emitted_path.exists()

    with emitted_path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.reader(handle))

    flattened_meta = excel_writer._flatten_dict(dict(processed[0].meta or {}))
    meta_items = sorted(flattened_meta.items())

    assert rows[0] == ["metadata_key", "metadata_value"]
    blank_row_index = rows.index([], 1)
    metadata_rows = rows[1:blank_row_index]
    metadata_map = {key: value for key, value in metadata_rows}
    expected_map = {
        key: "" if cleaned is None else str(cleaned)
        for key, cleaned in (
            (key, excel_writer._clean_value(value)) for key, value in meta_items
        )
    }
    assert metadata_map == expected_map
    assert rows[blank_row_index + 1][0] == "wavelength"

    report_results = plugin._report_context.get("results", {})
    per_spec_paths = report_results.get("per_spectrum_exports") or []
    assert str(emitted_path) in per_spec_paths
    assert any(str(emitted_path) in entry for entry in result.audit)
