from __future__ import annotations

import csv
import json
import math
import os
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple

import numpy as np
from openpyxl import Workbook

from spectro_app.engine.plugin_api import Spectrum


def _ensure_parent(path: Path) -> None:
    if not path.parent.exists():
        path.parent.mkdir(parents=True, exist_ok=True)


def _clean_value(value: Any) -> Any:
    if isinstance(value, np.ndarray):
        value = value.tolist()
    elif isinstance(value, np.generic):
        value = value.item()
    if isinstance(value, (np.floating, np.integer)):
        value = float(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        return value
    if isinstance(value, (list, tuple, set)):
        return json.dumps([_clean_value(v) for v in value])
    if isinstance(value, dict):
        return json.dumps({str(k): _clean_value(v) for k, v in value.items()})
    if isinstance(value, (Path, os.PathLike)):
        value = os.fspath(value)
    if isinstance(value, bytes):
        value = value.decode("utf-8", errors="replace")
    if isinstance(value, str):
        if value and value[0] in "=+-@":
            if not value.startswith("'"):
                return "'" + value
        return value
    return value


def _flatten_dict(data: Dict[str, Any], prefix: str = "") -> Dict[str, Any]:
    flat: Dict[str, Any] = {}
    for key, value in data.items():
        safe_key = str(key).replace(" ", "_").replace("/", "_")
        new_key = f"{prefix}{safe_key}" if not prefix else f"{prefix}.{safe_key}"
        if isinstance(value, dict):
            flat.update(_flatten_dict(value, new_key))
        elif isinstance(value, np.ndarray):
            flat[new_key] = _clean_value(value)
        elif isinstance(value, (list, tuple)):
            if value and all(isinstance(item, dict) for item in value):
                flat[new_key] = json.dumps([_clean_value(item) for item in value])
            else:
                flat[new_key] = json.dumps([_clean_value(item) for item in value])
        else:
            flat[new_key] = _clean_value(value)
    return flat


_STAGE_CHANNEL_ORDER: Tuple[str, ...] = (
    "raw",
    "blanked",
    "baseline_corrected",
    "joined",
    "despiked",
    "smoothed",
)


def _normalize_channel_label(value: Any) -> str:
    cleaned = _clean_value(value)
    if cleaned is None:
        return ""
    return str(cleaned)


def _format_channel_column(channel: str, wavelength: float) -> str:
    return f"{channel}:{wavelength:.6g}"


def _iter_channels(spec: Spectrum, wavelengths: np.ndarray) -> List[Tuple[str, np.ndarray]]:
    meta = spec.meta or {}
    channels: List[Tuple[str, np.ndarray]] = []
    base_label = _normalize_channel_label(meta.get("channel_label", "processed")) or "processed"
    intensity = np.asarray(spec.intensity, dtype=float)
    if intensity.shape == wavelengths.shape:
        channels.append((base_label, intensity))

    extra_channels = meta.get("channels") or {}
    ordered_names = [name for name in _STAGE_CHANNEL_ORDER if name in extra_channels]
    ordered_names.extend(name for name in extra_channels if name not in ordered_names)
    for name in ordered_names:
        arr = np.asarray(extra_channels[name], dtype=float)
        if arr.shape != wavelengths.shape:
            continue
        label = _normalize_channel_label(name) or str(name)
        channels.append((label, arr))
    return channels


def _processed_table_tidy(processed: Sequence[Spectrum]) -> Tuple[List[str], List[List[Any]]]:
    header = [
        "spectrum_index",
        "sample_id",
        "role",
        "mode",
        "channel",
        "wavelength",
        "value",
    ]
    rows: List[List[Any]] = []
    for idx, spec in enumerate(processed):
        wavelengths = np.asarray(spec.wavelength, dtype=float)
        meta = spec.meta or {}
        sample_id = _clean_value(
            meta.get("sample_id")
            or meta.get("channel")
            or meta.get("blank_id")
            or f"spec_{idx}"
        )
        role = _clean_value(meta.get("role", "sample"))
        mode = _clean_value(meta.get("mode"))
        for channel_label, data in _iter_channels(spec, wavelengths):
            for wl_val, inten_val in zip(wavelengths, data):
                rows.append(
                    [
                        idx,
                        sample_id,
                        role,
                        mode,
                        channel_label,
                        _clean_value(wl_val),
                        _clean_value(inten_val),
                    ]
                )
    return header, rows


def _processed_table_wide(processed: Sequence[Spectrum]) -> Tuple[List[str], List[List[Any]]]:
    if not processed:
        return ["wavelength"], []

    wavelength_grid: np.ndarray | None = None
    column_order: List[str] = []
    column_data: Dict[str, Dict[float, Any]] = {}

    for idx, spec in enumerate(processed):
        wavelengths = np.asarray(spec.wavelength, dtype=float).reshape(-1)
        if wavelength_grid is None:
            wavelength_grid = wavelengths
        else:
            if (
                wavelengths.shape != wavelength_grid.shape
                or not np.allclose(wavelengths, wavelength_grid, equal_nan=True)
            ):
                combined = np.concatenate((wavelength_grid, wavelengths))
                unique, indices = np.unique(combined, return_index=True)
                order = np.argsort(indices)
                wavelength_grid = unique[order]

        meta = spec.meta or {}
        sample_id_value = _clean_value(
            meta.get("sample_id")
            or meta.get("channel")
            or meta.get("blank_id")
            or f"spec_{idx}"
        )
        if sample_id_value in (None, ""):
            sample_id_value = f"spec_{idx}"
        sample_label = str(sample_id_value)

        for channel_label, data in _iter_channels(spec, wavelengths):
            channel_name = channel_label or "processed"
            column_name = f"{sample_label}:{channel_name}" if channel_name else sample_label
            if column_name in column_data:
                raise ValueError(
                    f"Duplicate column for sample '{sample_label}' and channel '{channel_name}'"
                )
            column_order.append(column_name)
            column_data[column_name] = {
                float(wl): _clean_value(inten_val)
                for wl, inten_val in zip(wavelengths, data)
            }

    assert wavelength_grid is not None  # for mypy
    header = ["wavelength"] + column_order
    rows: List[List[Any]] = []
    for wl in wavelength_grid:
        wl_value = float(wl)
        row = [_clean_value(wl_value)]
        for column_name in column_order:
            row.append(column_data[column_name].get(wl_value))
        rows.append(row)

    return header, rows


def _processed_table(processed: Sequence[Spectrum], layout: str) -> Tuple[List[str], List[List[Any]]]:
    if layout == "wide":
        return _processed_table_wide(processed)
    return _processed_table_tidy(processed)


def _metadata_rows(processed: Sequence[Spectrum]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for idx, spec in enumerate(processed):
        meta = dict(spec.meta or {})
        meta.pop("channels", None)
        flat = _flatten_dict(meta)
        flat["spectrum_index"] = idx
        rows.append(flat)
    return rows


def _qc_rows(qc_table: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    output: List[Dict[str, Any]] = []
    for row in qc_table:
        output.append(_flatten_dict(dict(row)))
    return output


def _write_dict_rows(ws, rows: List[Dict[str, Any]]):
    if not rows:
        return
    headers = sorted({key for row in rows for key in row.keys()})
    ws.append(headers)
    for row in rows:
        ws.append([_clean_value(row.get(header)) for header in headers])


def _write_calibration_sheet(ws, calibration: Dict[str, Any] | None) -> None:
    targets = calibration.get("targets", []) if calibration else []
    if not targets:
        ws.append(["Section", "Details"])
        ws.append(["Status", "No calibration performed"])
        return

    summary_header = [
        "Target",
        "Status",
        "Slope",
        "Intercept",
        "R^2",
        "Residual Std",
        "LOD",
        "LOQ",
        "Points",
    ]
    ws.append(summary_header)
    for target in targets:
        fit = target.get("fit") or {}
        ws.append(
            [
                _clean_value(target.get("name")),
                _clean_value(target.get("status")),
                _clean_value(fit.get("slope")),
                _clean_value(fit.get("intercept")),
                _clean_value(fit.get("r_squared")),
                _clean_value(fit.get("residual_std")),
                _clean_value(fit.get("lod")),
                _clean_value(fit.get("loq")),
                _clean_value(fit.get("points")),
            ]
        )

    for target in targets:
        ws.append([])
        ws.append([f"Target: {target.get('name')}"])
        for error in target.get("errors", []):
            ws.append(["Error", _clean_value(error)])
        for warning in target.get("warnings", []):
            ws.append(["Warning", _clean_value(warning)])

        standards = target.get("standards") or []
        if standards:
            ws.append(["Standards"])
            std_header = [
                "Sample ID",
                "Concentration",
                "Response (A/cm)",
                "Raw Absorbance",
                "Predicted Response",
                "Residual",
                "Included",
                "Replicates",
                "Pathlength (cm)",
                "Configured Weight",
                "Applied Weight",
            ]
            ws.append(std_header)
            for entry in standards:
                ws.append(
                    [
                        _clean_value(entry.get("sample_id")),
                        _clean_value(entry.get("concentration")),
                        _clean_value(entry.get("response")),
                        _clean_value(entry.get("raw_absorbance")),
                        _clean_value(entry.get("predicted_response")),
                        _clean_value(entry.get("residual")),
                        _clean_value(entry.get("included")),
                        _clean_value(entry.get("replicates")),
                        _clean_value(entry.get("pathlength_cm")),
                        _clean_value(entry.get("weight")),
                        _clean_value(entry.get("applied_weight")),
                    ]
                )
        else:
            ws.append(["Standards", "None"])

        unknowns = target.get("unknowns") or []
        if unknowns:
            ws.append(["Unknowns"])
            unk_header = [
                "Sample ID",
                "Response (A/cm)",
                "Raw Absorbance",
                "Predicted Concentration",
                "Predicted Concentration StdErr",
                "Status",
                "Within Range",
                "LOD/LOQ Flag",
                "Replicates",
                "Pathlength (cm)",
                "Dilution Factor",
            ]
            ws.append(unk_header)
            for entry in unknowns:
                ws.append(
                    [
                        _clean_value(entry.get("sample_id")),
                        _clean_value(entry.get("response")),
                        _clean_value(entry.get("raw_absorbance")),
                        _clean_value(entry.get("predicted_concentration")),
                        _clean_value(entry.get("predicted_concentration_std_err")),
                        _clean_value(entry.get("status")),
                        _clean_value(entry.get("within_range")),
                        _clean_value(entry.get("lod_flag")),
                        _clean_value(entry.get("replicates")),
                        _clean_value(entry.get("pathlength_cm")),
                        _clean_value(entry.get("dilution_factor")),
                    ]
                )
        else:
            ws.append(["Unknowns", "None"])


def write_workbook(
    out_path: str,
    processed,
    qc_table,
    audit,
    figures,
    calibration=None,
    *,
    processed_layout: str = "tidy",
):
    workbook_path = Path(out_path)
    _ensure_parent(workbook_path)

    wb = Workbook()
    ws_processed = wb.active
    ws_processed.title = "Processed_Spectra"
    header, rows = _processed_table(processed, processed_layout)
    ws_processed.append(header)
    for row in rows:
        ws_processed.append(row)

    ws_metadata = wb.create_sheet("Metadata")
    _write_dict_rows(ws_metadata, _metadata_rows(processed))

    ws_qc = wb.create_sheet("QC_Flags")
    _write_dict_rows(ws_qc, _qc_rows(qc_table))

    ws_calibration = wb.create_sheet("Calibration")
    _write_calibration_sheet(ws_calibration, calibration)

    ws_audit = wb.create_sheet("Audit_Log")
    ws_audit.append(["Index", "Entry"])
    for idx, entry in enumerate(audit or [], start=1):
        ws_audit.append([idx, _clean_value(entry)])

    wb.save(workbook_path)
    return str(workbook_path)


def write_single_spectrum_workbook(out_path: str | Path, spectrum: Spectrum) -> Path:
    """Write ``spectrum`` to a single-sheet workbook and return the path."""

    workbook_path = Path(out_path)
    _ensure_parent(workbook_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Spectrum"

    wavelengths = np.asarray(spectrum.wavelength, dtype=float)
    channels = _iter_channels(spectrum, wavelengths)
    header = ["wavelength"] + [label for label, _ in channels]
    ws.append(header)
    for idx, wavelength in enumerate(wavelengths):
        row = [_clean_value(wavelength)]
        for _, channel_data in channels:
            if idx < channel_data.shape[0]:
                row.append(_clean_value(channel_data[idx]))
            else:
                row.append(None)
        ws.append(row)

    metadata = wb.create_sheet("Metadata")
    metadata.append(["key", "value"])
    for key, value in sorted(_flatten_dict(dict(spectrum.meta or {})).items()):
        metadata.append([key, _clean_value(value)])

    wb.save(workbook_path)
    return workbook_path


def write_single_spectrum_csv(out_path: str | Path, spectrum: Spectrum) -> Path:
    """Write ``spectrum`` to a CSV file and return the emitted path."""

    csv_path = Path(out_path)
    _ensure_parent(csv_path)

    wavelengths = np.asarray(spectrum.wavelength, dtype=float)
    channels = _iter_channels(spectrum, wavelengths)
    header = ["wavelength"] + [label for label, _ in channels]

    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)

        flattened_meta = _flatten_dict(dict(spectrum.meta or {}))
        writer.writerow(["metadata_key", "metadata_value"])
        for key, value in sorted(flattened_meta.items()):
            writer.writerow([key, _clean_value(value)])

        writer.writerow([])
        writer.writerow(header)
        for idx, wavelength in enumerate(wavelengths):
            row = [_clean_value(wavelength)]
            for _, channel_data in channels:
                if idx < channel_data.shape[0]:
                    row.append(_clean_value(channel_data[idx]))
                else:
                    row.append(None)
            writer.writerow(row)

    return csv_path
